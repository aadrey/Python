from openpyxl import  load_workbook

wb1 = load_workbook('d:\Afonin\Python\Cable\Sheet1\Cable.xlsx')
print(wb1.sheetnames)

sheet1 = wb1 ['Кабельный журнал']
sheet2 = wb1 ['Щит']

# Функция определения разницы координат между двух точек на чертеже и записи его в файл в текущую строку кабельного журнала
def raznica_koord (numcell_1,numcell_2):
    raznica_x = abs(float (sheet2 ['D'+numcell_1].value) - float (sheet2 ['D'+numcell_2].value))
    print ('Длина X=',raznica_x)
    raznica_y = abs(float (sheet2 ['E'+numcell_1].value) - float (sheet2 ['E'+numcell_2].value))
    print ('Длина Y=',raznica_y)
    raznica = raznica_x + raznica_y
    print('Длина всего ' + numcell_1 + '-' + numcell_2 + ' в строке ' + str(number_str) + ' =', raznica)
    sheet1['H' + str(number_str)] = raznica
    wb1.save('d:\Afonin\Python\Cable\Sheet1\Cable.xlsx')

# Цикл по строкам кабельного журнала
for number_str in range(1,16):

    # Определение точки 1 (откуда идёт кабель)
    point_1 = sheet1 ['C'+str(number_str)].value
    if "X" in point_1:
        point_1 = sheet1 ['A1'].value
    print ('Точка 1 =',point_1)

    # Определение точки 2 (куда идёт кабель)
    point_2 = sheet1 ['D'+str(number_str)].value
    if "X" in point_2:
        point_2 = sheet1 ['A1'].value
    print ('Точка 2 =',point_2)

    # Определение строки с координатами для точки 1
    stroka_1 = 2
    while point_1 != sheet2 ['C'+str(stroka_1)].value:
        stroka_1 = stroka_1 + 1
        if stroka_1 == 16:
            break
    print (stroka_1)

    # Определение строки с координатами для точки 2
    stroka_2 = 2
    while point_2 != sheet2 ['C'+str(stroka_2)].value:
        stroka_2 = stroka_2 + 1
        if stroka_2 == 16:
            break
    print (stroka_2)

    # Вызов функции для определения расстояния между точками 1 и 2 для текущей строки кабельного журнала
    cell_1 = str(stroka_1)
    cell_2 = str(stroka_2)
    raznica_koord (cell_1,cell_2)